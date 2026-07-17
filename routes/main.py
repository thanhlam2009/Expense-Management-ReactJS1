from flask import Blueprint, jsonify, make_response
from flask_login import login_required, current_user
from models.transaction import Transaction
from app import db
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

main_bp = Blueprint('main', __name__)

@main_bp.route('/')
def index():
    """Health-check của Backend API
    ---
    tags:
      - Export
    responses:
      200:
        description: Backend đang hoạt động
    """
    return jsonify({'message': 'Expense management API'})

@main_bp.route('/export/transactions')
@login_required
def export_transactions():
    """Xuất toàn bộ giao dịch ra file Excel (.xlsx)
    ---
    tags:
      - Export
    produces:
      - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    responses:
      200:
        description: File Excel (.xlsx) chứa toàn bộ giao dịch kèm tổng kết
    """
    try:
        # Lấy tất cả giao dịch của user
        transactions = Transaction.query.filter_by(user_id=current_user.id)\
                                      .order_by(Transaction.date.desc())\
                                      .all()
        
        # Tạo workbook và worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Giao Dịch"
        
        # Định nghĩa headers
        headers = [
            'STT', 'Ngày', 'Loại', 'Danh mục', 'Mô tả', 
            'Số tiền (VNĐ)', 'Hóa đơn', 'Ngày tạo'
        ]
        
        # Style cho header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Thêm headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Thêm dữ liệu
        for row, transaction in enumerate(transactions, 2):
            # STT
            ws.cell(row=row, column=1, value=row-1).border = border
            
            # Ngày
            date_cell = ws.cell(row=row, column=2, value=transaction.date.strftime('%d/%m/%Y'))
            date_cell.border = border
            
            # Loại
            type_text = "Thu nhập" if transaction.type == 'income' else "Chi tiêu"
            type_cell = ws.cell(row=row, column=3, value=type_text)
            type_cell.border = border
            # Màu sắc theo loại
            if transaction.type == 'income':
                type_cell.font = Font(color="008000")  # Xanh lá
            else:
                type_cell.font = Font(color="FF0000")  # Đỏ
            
            # Danh mục
            category_cell = ws.cell(row=row, column=4, value=transaction.category.name if transaction.category else "N/A")
            category_cell.border = border
            
            # Mô tả
            desc_cell = ws.cell(row=row, column=5, value=transaction.description or "")
            desc_cell.border = border
            
            # Số tiền
            amount_cell = ws.cell(row=row, column=6, value=float(transaction.amount))
            amount_cell.number_format = '#,##0'
            amount_cell.border = border
            amount_cell.alignment = Alignment(horizontal="right")
            
            # Hóa đơn
            receipt_cell = ws.cell(row=row, column=7, value="Có" if transaction.receipt_image else "Không")
            receipt_cell.border = border
            
            # Ngày tạo
            created_cell = ws.cell(row=row, column=8, value=transaction.created_at.strftime('%d/%m/%Y %H:%M'))
            created_cell.border = border
        
        # Tự động điều chỉnh độ rộng cột
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Thêm thống kê ở cuối
        if transactions:
            stats_row = len(transactions) + 3
            
            # Tổng thu nhập
            total_income = sum(t.amount for t in transactions if t.type == 'income')
            ws.cell(row=stats_row, column=5, value="Tổng thu nhập:").font = Font(bold=True)
            income_cell = ws.cell(row=stats_row, column=6, value=float(total_income))
            income_cell.number_format = '#,##0'
            income_cell.font = Font(bold=True, color="008000")
            
            # Tổng chi tiêu
            total_expense = sum(t.amount for t in transactions if t.type == 'expense')
            ws.cell(row=stats_row + 1, column=5, value="Tổng chi tiêu:").font = Font(bold=True)
            expense_cell = ws.cell(row=stats_row + 1, column=6, value=float(total_expense))
            expense_cell.number_format = '#,##0'
            expense_cell.font = Font(bold=True, color="FF0000")
            
            # Số dư
            balance = total_income - total_expense
            ws.cell(row=stats_row + 2, column=5, value="Số dư:").font = Font(bold=True)
            balance_cell = ws.cell(row=stats_row + 2, column=6, value=float(balance))
            balance_cell.number_format = '#,##0'
            balance_cell.font = Font(bold=True, color="0000FF")
        
        # Lưu file vào memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Tạo tên file với timestamp
        filename = f"giao_dich_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Tạo response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({'error': f'Lỗi khi export Excel: {str(e)}'}), 500